import os
import secrets
from datetime import datetime, timedelta

import pymysql
from flask import Flask, jsonify, redirect, render_template, request, session, url_for
from openpyxl import load_workbook

app = Flask(__name__, static_folder="assets", static_url_path="/assets", template_folder="templates")
app.secret_key = os.getenv("APP_SECRET", "change_me")


def db():
    return pymysql.connect(
        host=os.getenv("DB_HOST", "db"),
        user=os.getenv("DB_USER", "parking"),
        password=os.getenv("DB_PASSWORD", "parking_pass"),
        database=os.getenv("DB_NAME", "parking"),
        port=int(os.getenv("DB_PORT", "3306")),
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
    )


def normalize_phone(phone: str) -> str:
    digits = "".join(ch for ch in phone if ch.isdigit())
    if not digits:
        return ""
    if len(digits) == 10:
        return "7" + digits
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        return "7" + digits[1:]
    return digits


def normalize_plate(text: str) -> str:
    clean = text.replace(" ", "").replace("-", "")
    mapping = {
        "а": "A",
        "в": "B",
        "е": "E",
        "к": "K",
        "м": "M",
        "н": "H",
        "о": "O",
        "р": "P",
        "с": "C",
        "т": "T",
        "у": "Y",
        "х": "X",
        "А": "A",
        "В": "B",
        "Е": "E",
        "К": "K",
        "М": "M",
        "Н": "H",
        "О": "O",
        "Р": "P",
        "С": "C",
        "Т": "T",
        "У": "Y",
        "Х": "X",
    }
    normalized = "".join(mapping.get(ch, ch) for ch in clean)
    return normalized.upper()


def role_label(role: str) -> str:
    return {
        "admin": "Администратор",
        "manager": "Менеджер",
        "guard": "Охранник",
    }.get(role, role)


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    with db().cursor() as cur:
        cur.execute("SELECT * FROM users WHERE id=%s", (user_id,))
        return cur.fetchone()


def require_login():
    user = current_user()
    if not user:
        return redirect(url_for("login"))
    return user


def csrf_token():
    if "csrf" not in session:
        session["csrf"] = secrets.token_hex(16)
    return session["csrf"]


def csrf_validate():
    if request.method == "POST":
        if request.form.get("csrf_token") != session.get("csrf"):
            return False
    return True


def ensure_admin():
    phone = normalize_phone(os.getenv("ADM_PHONE", ""))
    if not phone:
        return
    name = os.getenv("ADM_NAME", "Administrator")
    with db().cursor() as cur:
        cur.execute("SELECT id, role, status FROM users WHERE phone=%s", (phone,))
        user = cur.fetchone()
        if user:
            if user["role"] != "admin" or user["status"] != "active":
                cur.execute("UPDATE users SET role='admin', status='active' WHERE id=%s", (user["id"],))
            return
        cur.execute(
            "INSERT INTO users (name, phone, role, status) VALUES (%s, %s, 'admin', 'active')",
            (name, phone),
        )


@app.before_request
def bootstrap():
    ensure_admin()


@app.context_processor
def inject_globals():
    return {"role_label": role_label, "csrf_token": csrf_token}


@app.route("/")
def index():
    user = require_login()
    if not isinstance(user, dict):
        return user
    if user["role"] == "admin":
        return redirect(url_for("users"))
    return redirect(url_for("cars"))


@app.route("/login", methods=["GET", "POST"])
@app.route("/login.php", methods=["GET", "POST"])
def login():
    message = ""
    if request.method == "POST":
        if not csrf_validate():
            message = "Ошибка CSRF"
        else:
            phone = normalize_phone(request.form.get("phone", ""))
            if not phone:
                message = "Введите корректный номер телефона."
            else:
                with db().cursor() as cur:
                    cur.execute(
                        "SELECT * FROM users WHERE phone=%s AND status='active'",
                        (phone,),
                    )
                    user = cur.fetchone()
                    if not user:
                        message = "Пользователь не найден или заблокирован."
                    elif not user.get("tg_id"):
                        message = "Привяжите Telegram: напишите боту /start и отправьте телефон."
                    else:
                        code = "".join(secrets.choice("ABCDEFGHJKLMNPQRSTUVWXYZ23456789") for _ in range(8))
                        code_hash = secrets.token_hex(16)
                        cur.execute(
                            "INSERT INTO otp_sessions (user_id, code_hash, expires_at) VALUES (%s, %s, DATE_ADD(NOW(), INTERVAL 60 SECOND))",
                            (user["id"], code_hash),
                        )
                        cur.execute(
                            "INSERT INTO otp_outbox (user_id, message) VALUES (%s, %s)",
                            (user["id"], f"<code>{code}</code>"),
                        )
                        session["pending_user_id"] = user["id"]
                        return redirect(url_for("otp"))
    return render_template("login.html", message=message, user=None, current="")


@app.route("/otp", methods=["GET", "POST"])
@app.route("/otp.php", methods=["GET", "POST"])
def otp():
    if "pending_user_id" not in session:
        return redirect(url_for("login"))
    message = ""
    if request.method == "POST":
        if not csrf_validate():
            message = "Ошибка CSRF"
        else:
            code = request.form.get("code", "").strip().upper()
            with db().cursor() as cur:
                cur.execute(
                    "SELECT * FROM otp_sessions WHERE user_id=%s ORDER BY id DESC LIMIT 1",
                    (session["pending_user_id"],),
                )
                otp_session = cur.fetchone()
                if not otp_session:
                    message = "Неверный или просроченный OTP."
                else:
                    expires = otp_session["expires_at"]
                    if expires < datetime.now():
                        message = "Неверный или просроченный OTP."
                    else:
                        session["user_id"] = session["pending_user_id"]
                        session.pop("pending_user_id", None)
                        return redirect(url_for("index"))
    return render_template("otp.html", message=message, user=None, current="")


@app.route("/logout")
@app.route("/logout.php")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/cars")
@app.route("/cars.php")
def cars():
    user = require_login()
    if not isinstance(user, dict):
        return user
    page = max(int(request.args.get("page", 1)), 1)
    limit = 20
    offset = (page - 1) * limit
    with db().cursor() as cur:
        cur.execute("SELECT COUNT(*) as cnt FROM cars")
        total = cur.fetchone()["cnt"]
        if user["role"] == "admin":
            cur.execute(
                "SELECT cars.*, users.name AS added_by_name, users.phone AS added_by_phone "
                "FROM cars LEFT JOIN users ON cars.who_added = users.id "
                "ORDER BY date_added DESC LIMIT %s OFFSET %s",
                (limit, offset),
            )
        else:
            cur.execute(
                "SELECT * FROM cars ORDER BY date_added DESC LIMIT %s OFFSET %s",
                (limit, offset),
            )
        rows = cur.fetchall()
    pages = max(1, (total + limit - 1) // limit)
    return render_template("cars.html", user=user, cars=rows, page=page, pages=pages, current="cars")


@app.route("/users")
@app.route("/users.php")
def users():
    user = require_login()
    if not isinstance(user, dict):
        return user
    if user["role"] not in {"admin", "manager"}:
        return redirect(url_for("cars"))
    with db().cursor() as cur:
        if user["role"] == "manager":
            cur.execute("SELECT * FROM users WHERE role != 'admin' ORDER BY created_at DESC")
        else:
            cur.execute("SELECT * FROM users ORDER BY created_at DESC")
        rows = cur.fetchall()
    return render_template("users.html", user=user, users=rows, current="users")


@app.route("/remote_cars")
@app.route("/remote_cars.php")
def remote_cars():
    user = require_login()
    if not isinstance(user, dict):
        return user
    if user["role"] != "admin":
        return redirect(url_for("cars"))
    with db().cursor() as cur:
        cur.execute(
            "SELECT remote_cars.*, added.name AS added_by_name, added.phone AS added_by_phone, "
            "deleted.name AS deleted_by_name, deleted.phone AS deleted_by_phone "
            "FROM remote_cars "
            "LEFT JOIN users AS added ON remote_cars.who_added = added.id "
            "LEFT JOIN users AS deleted ON remote_cars.who_deleted = deleted.id "
            "ORDER BY date_deleted DESC"
        )
        rows = cur.fetchall()
    return render_template("remote_cars.html", user=user, rows=rows, current="remote")


@app.route("/api/cars_search.php")
@app.route("/api/cars_search")
def api_cars_search():
    user = require_login()
    if not isinstance(user, dict):
        return user
    q = normalize_plate(request.args.get("q", ""))
    if not q:
        return jsonify({"results": []})
    expr = (
        "REPLACE(REPLACE(UPPER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(car_number, ' ', ''), '-', ''), 'А', 'A'), 'В', 'B'), 'Е', 'E'), 'К', 'K'), 'М', 'M'), 'Н', 'H'), 'О', 'O'), 'Р', 'P'), 'С', 'C'), 'Т', 'T'), 'У', 'Y'), 'Х', 'X'))"
    )
    with db().cursor() as cur:
        cur.execute(
            f"SELECT car_number, car_model, comment, date_added FROM cars WHERE {expr} LIKE %s ORDER BY date_added DESC LIMIT 20",
            (f"%{q}%",),
        )
        rows = cur.fetchall()
    return jsonify({"results": rows})


@app.route("/api/users", methods=["POST"])
@app.route("/api/users_crud.php", methods=["POST"])
def users_crud():
    user = require_login()
    if not isinstance(user, dict):
        return user
    if user["role"] not in {"admin", "manager"}:
        return redirect(url_for("users"))
    if not csrf_validate():
        return redirect(url_for("users"))
    action = request.form.get("action", "")
    with db().cursor() as cur:
        if action == "create":
            name = request.form.get("name", "").strip()
            phone = normalize_phone(request.form.get("phone", ""))
            role = request.form.get("role", "guard")
            if user["role"] == "manager" and role == "admin":
                return redirect(url_for("users"))
            if name and phone:
                cur.execute(
                    "INSERT INTO users (name, phone, role, status) VALUES (%s, %s, %s, 'active')",
                    (name, phone, role),
                )
        elif action == "toggle_status":
            target_id = int(request.form.get("user_id", 0))
            cur.execute("SELECT id, role, status FROM users WHERE id=%s", (target_id,))
            target = cur.fetchone()
            if target and not (user["role"] == "manager" and target["role"] == "admin"):
                new_status = "blocked" if target["status"] == "active" else "active"
                cur.execute("UPDATE users SET status=%s WHERE id=%s", (new_status, target_id))
        elif action == "delete":
            target_id = int(request.form.get("user_id", 0))
            cur.execute("SELECT id, role FROM users WHERE id=%s", (target_id,))
            target = cur.fetchone()
            if target and not (user["role"] == "manager" and target["role"] == "admin"):
                cur.execute("DELETE FROM users WHERE id=%s", (target_id,))
    return redirect(url_for("users"))


@app.route("/api/cars_crud.php", methods=["POST"])
def cars_crud():
    user = require_login()
    if not isinstance(user, dict):
        return user
    if user["role"] not in {"admin", "manager"}:
        return redirect(url_for("cars"))
    if not csrf_validate():
        return redirect(url_for("cars"))
    action = request.form.get("action", "")
    with db().cursor() as cur:
        if action == "create":
            model = request.form.get("car_model", "").strip()
            number = normalize_plate(request.form.get("car_number", ""))
            comment = request.form.get("comment", "").strip()
            if model and number:
                cur.execute(
                    "INSERT INTO cars (car_model, car_number, comment, who_added, date_added) VALUES (%s, %s, %s, %s, NOW())",
                    (model, number, comment, user["id"]),
                )
        elif action == "delete":
            car_id = int(request.form.get("car_id", 0))
            cur.execute("SELECT * FROM cars WHERE id=%s", (car_id,))
            car = cur.fetchone()
            if car:
                cur.execute(
                    "INSERT INTO remote_cars (car_model, car_number, comment, who_added, date_added, who_deleted, date_deleted) VALUES (%s, %s, %s, %s, %s, %s, NOW())",
                    (car["car_model"], car["car_number"], car["comment"], car["who_added"], car["date_added"], user["id"]),
                )
                cur.execute("DELETE FROM cars WHERE id=%s", (car_id,))
        elif action == "import":
            file = request.files.get("import_file")
            if file and file.filename:
                workbook = load_workbook(file)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    if not row or len(row) < 2:
                        continue
                    model = (row[0] or "").strip()
                    number = normalize_plate(row[1] or "")
                    comment = (row[2] or "").strip() if len(row) > 2 else ""
                    if model and number:
                        cur.execute(
                            "INSERT INTO cars (car_model, car_number, comment, who_added, date_added) VALUES (%s, %s, %s, %s, NOW())",
                            (model, number, comment, user["id"]),
                        )
    return redirect(url_for("cars"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=80)
